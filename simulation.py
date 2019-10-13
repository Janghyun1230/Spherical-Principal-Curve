from principal_curves import *
import time
import numpy as np
import openpyxl


def simulated_data(d_type, sigma1=0.0, sigma2=0.07, n=100, is_print=True, freq=8, amp=1 / 3, wind=1.5):
    # data generation
    if d_type == "circle":
        theta = np.linspace(0, np.pi * 2, n) + np.pi
        phi = np.array([np.pi / 4] * n)
    elif d_type == "wave":
        theta = np.linspace(0, np.pi * 2, n)
        phi = np.sin(theta * freq) * amp + np.pi / 2
    elif d_type == "spiral":
        theta = np.linspace(0, np.pi * 2 * wind, n) + np.pi
        phi = np.linspace(np.pi / 8, np.pi / 2, n)
    else:
        raise AssertionError("Wrong d_type !! d_type should be one of (circle, wave, spiral)")

    x_true = np.vstack((np.sin(phi) * np.cos(theta), np.sin(phi) * np.sin(theta), np.cos(phi)))

    noise1 = np.random.normal(0, sigma1, n)
    noise2 = np.random.normal(0, sigma2, n)
    theta = theta + noise1
    phi = phi + noise2
    x_noise = np.vstack((np.sin(phi) * np.cos(theta), np.sin(phi) * np.sin(theta), np.cos(phi)))

    x_polar = np.vstack((phi, theta))

    if is_print:
        plot_sphere(x_noise, title="{} (sigma : {})".format(d_type, sigma2), extra=x_true, line=True)

    return x_true, x_noise, x_polar


def simulation(d_type="circle", q=0.05, n_point=100, sigma2_list=[0.07, 0.1], repeat=50):
    wb = openpyxl.Workbook()

    for j, sigma2 in enumerate(sigma2_list):
        print("\nsigma : {} is start !!".format(sigma2))

        ws = wb.create_sheet("sigma {}".format(sigma2))
        ws.cell(row= 1,column= 1).value = 'repeat'
        ws.cell(row= 1,column= 2).value = 'Extrinsic'

        ws.cell(row= 2,column= 2).value = 'noisy recon'
        ws.cell(row= 2,column= 3).value = 'noisy projections'
        ws.cell(row= 2,column= 4).value = 'clean recon'
        ws.cell(row= 2,column= 5).value = 'clean projections'

        ws.cell(row= 1,column= 7).value = 'Intrinsic'

        ws.cell(row= 2,column= 7).value = 'noisy recon'
        ws.cell(row= 2,column= 8).value = 'noisy projections'
        ws.cell(row= 2,column= 9).value = 'clean recon'
        ws.cell(row= 2,column= 10).value = 'clean projections'

        ws.cell(row= 1,column= 12).value = 'Hauberg'

        ws.cell(row= 2,column= 12).value = 'noisy recon'
        ws.cell(row= 2,column= 13).value = 'noisy projections'
        ws.cell(row= 2,column= 14).value = 'clean recon'
        ws.cell(row= 2,column= 15).value = 'clean projections'

        for n in range(repeat):
            print("{} / {} processing ...".format(n+1, repeat))
            if n ==0 :
                x_true, x_noise, x_polar = simulated_data(d_type=d_type, sigma2=sigma2, is_print=False)
            else :
                x_true, x_noise, x_polar = simulated_data(d_type=d_type, sigma2=sigma2, is_print=False)

            # initialization (principal circle)
            p_, p3 = pga_circle(x_polar, False)

            # initial projection
            init_project = np.zeros(shape=[2, n_point])
            init_project[0, :] = p3
            init_project[1, :] = np.arange(0, 2*m.pi, 2*m.pi/n_point)
            init_project = inv_rotate(p_, rad2euclid(init_project))

            # exact (extrinsic)
            points, train_dist, lambda_arr= train_principal_curve(data=x_noise, init_project=init_project, iter=10,
                                                                  neigh_ratio=q, intrinsic=False, exact=True, is_print=False)
            proj_num_train = len(np.unique(lambda_arr))

            d, points, lambda_arr, _, _= project_curve(x_true, points, exact = True)
            test_dist = sum(d**2)
            proj_num_test = len(np.unique(lambda_arr))

            ws.cell(row= n+3,column= 1).value = n+1
            ws.cell(row= n+3,column= 2).value = train_dist
            ws.cell(row= n+3,column= 3).value = proj_num_train
            ws.cell(row= n+3,column= 4).value = test_dist
            ws.cell(row= n+3,column= 5).value = proj_num_test

            # exact (intrinsic)
            points, train_dist, lambda_arr= train_principal_curve(data=x_noise, init_project=init_project, iter=10,
                                                                  neigh_ratio=q, intrinsic=True, exact=True, is_print=False)
            proj_num_train = len(np.unique(lambda_arr))

            d, points, lambda_arr, _, _= project_curve(x_true, points, exact = True)
            test_dist = sum(d**2)
            proj_num_test = len(np.unique(lambda_arr))

            ws.cell(row= n+3,column= 7).value = train_dist
            ws.cell(row= n+3,column= 8).value = proj_num_train
            ws.cell(row= n+3,column= 9).value = test_dist
            ws.cell(row= n+3,column= 10).value = proj_num_test

            # not exact (Haubergs)
            points, train_dist, lambda_arr= train_principal_curve(x_noise, init_project, iter=10,
                                                                  neigh_ratio=q, intrinsic=False, exact=False, is_print=False)
            proj_num_train = len(np.unique(lambda_arr))

            d, points, lambda_arr, _, _= project_curve(x_true, points, exact = False)
            test_dist = sum(d**2)
            proj_num_test = len(np.unique(lambda_arr))

            # save as excel file
            ws.cell(row= n+3,column= 12).value = train_dist
            ws.cell(row= n+3,column= 13).value = proj_num_train
            ws.cell(row= n+3,column= 14).value = test_dist
            ws.cell(row= n+3,column= 15).value = proj_num_train

            wb.save("results/{}_points{}_neigh{}.xlsx".format(d_type, n_point, q))

    print("file is saved as results/{}_points{}_neigh{}.xlsx ! please check excel sheet".format(d_type, d_type, n_point, q))
    wb.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--q', type=float)
    parser.add_argument('--n', type=int)

    args = parser.parse_args()

    simulation(d_type="circle", q=args.q, n_point=args.n)
    simulation(d_type="wave", q=args.q, n_point=args.n)

    plt.show()